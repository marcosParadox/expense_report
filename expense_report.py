import re, os
from pdfminer.high_level import extract_text
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

#Read the contents of the PDF
text = extract_text('statement.pdf')

#Transacton Date
# RegEx to match the pattern
dateRegex = re.compile(r'(\d+/\d+)')
#regex to remove due date from 1st page and date from footer last page
dateRemoveYearRegex = re.compile(r'(\d\d/\d\d/\d\d)')

# string to hold extracted text minus
new_string = re.sub(dateRemoveYearRegex, '', text)

#Find all matches for trans date
dateText = dateRegex.findall(new_string)
# Sort the date list (required for multi-page statements)
dateText.sort()
# Select every other date (removes post date)
dateText = dateText[::2]

#Transaction Amount
#Regex to remove $0.00 from account summary
amtRemoveRegex = re.compile(r'\$0\.00')
new_string = re.sub(amtRemoveRegex, '', new_string)
amtRegex = re.compile(r'-?\$\d{0,3}\.\d\d')
amtText = amtRegex.findall(new_string)

#Open destination workbook
wb = load_workbook('Expense Report template MP.xlsx')
ws = wb['Expense Report']

#Insert blank rows if needed
numberOfRows = len(amtText) 
maxRow = 23
lastRow = (ws.max_row)
if numberOfRows > maxRow:
    rowsToInsert = numberOfRows - maxRow
    ws.insert_rows(8,rowsToInsert)

# Insert dates
date_row = 8
for date in dateText:
	ws.cell(row=date_row, column=1).value = date
	date_row += 1

# Remove the Dollar Sign from the amtText
out_list = [re.sub(r'\$','',string) for string in amtText]

# Convert amounts into floating point numbers
amtFloat = []
for element in out_list:
    amtFloat.append(float(element))
# Insert Amounts
amt_row = 8
for amt in amtFloat:
	ws.cell(row=amt_row, column=3).value = amt
	amt_row += 1

#insert total formula in column G
formula_row = 8
for forms in amtFloat:
	ws.cell(row=formula_row, column=7).value = '=Sum(C' + str(formula_row) + ':C' + str(formula_row) + ')'
	formula_row += 1


#add the total sum for the expense report into the last row columns c,d,e,f,g. Current formula => =SUM(G8:G30)

ws.cell(row=formula_row, column=7).value = '=Sum(G8:G' + str(formula_row-1) + ')'
ws.cell(row=formula_row, column=6).value = '=Sum(F8:F' + str(formula_row-1) + ')'
ws.cell(row=formula_row, column=5).value = '=Sum(E8:E' + str(formula_row-1) + ')'
ws.cell(row=formula_row, column=4).value = '=Sum(D8:D' + str(formula_row-1) + ')'
ws.cell(row=formula_row, column=3).value = '=Sum(C8:C' + str(formula_row-1) + ')'
#Save the new workbook
wb.save('names.xlsx')
